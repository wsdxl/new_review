"""
============================
Author  : XiaoLei.Du
Time    : 2020/8/12 20:41
E-mail  : 506615839@qq.com
File    : 2020_08_12.py
============================
"""
import openpyxl
# 1、打开工作簿对象
work_book=openpyxl.load_workbook(r'F:\project\new_review\cases.xlsx')
# 2、选择表单
sheet=work_book['login']
# print(sheet)
# 3、选择第二行第二列的格子
data=sheet.cell(row=2,column=2)
# print(data.value)

# 写入格子
sheet.cell(row=7,column=2,value='python666')
# 写入后一定要保存才生效
work_book.save(r'F:\project\new_review\cases.xlsx')
# # 获取最大行
# print(sheet.max_row)
# # 获取最大列
# print(sheet.max_column)

# rows：按行获取所有的格子对象，每一行存在一个元祖中
print(list(sheet.rows))